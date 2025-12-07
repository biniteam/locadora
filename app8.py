import streamlit as st
import sqlite3
import pandas as pd
from datetime import date, datetime, timedelta
import os
#import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io # Importar io para lidar com dados em memória
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from pdfgenerator import gerar_contrato_pdf, gerar_recibo_pdf # Importa as funções do novo módulo
from pdfgenerator import STATUS_CARRO, STATUS_CLIENTE # Importa os status do novo módulo
from database_backup import interface_backup, fazer_backup # Importa funções de backup
from init_db import init_db_production, check_db_health # Importa inicialização robusta
from auth import auth_manager, login_page, logout, require_login, get_current_user, check_permission, USER_ROLES # Importa sistema de autenticação
import numpy as np


# --- INICIALIZAÇÃO DO BANCO DE DADOS ---
# Verificar e inicializar banco para produção
db_health = check_db_health()
if not db_health['healthy']:
    init_db_production()

# --- AUTENTICAÇÃO ---
# Verificar login antes de mostrar aplicação
if not require_login():
    st.stop()  # Para execução se não estiver logado

# Obter usuário atual
current_user = get_current_user()

# --- FUNÇÕES DE FORMATAÇÃO E UTILIDADE ---

def formatar_moeda(valor):
    """Formata um valor float para a moeda brasileira (R$ 0.000,00)."""
    if valor is None:
        valor = 0.0
    # Garante que o separador decimal seja vírgula e o milhar seja ponto
    return f"R$ {valor:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


# --- BANCO DE DADOS ---
def init_db():
    conn = sqlite3.connect('locadora_v2.db', detect_types=sqlite3.PARSE_DECLTYPES)
    c = conn.cursor()

    c.execute('''
        CREATE TABLE IF NOT EXISTS carros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            modelo TEXT,
            placa TEXT UNIQUE,
            cor TEXT,
            diaria REAL,
            preco_km REAL,
            km_atual INTEGER,
            status TEXT DEFAULT 'Disponível'  
        )
    ''')

    # Adiciona a coluna 'numero_chassi' na tabela carros se ela não existir
    c.execute('''
        PRAGMA table_info(carros);
    ''')
    columns_carros = [col[1] for col in c.fetchall()]
    if 'numero_chassi' not in columns_carros:
        c.execute('''
            ALTER TABLE carros ADD COLUMN numero_chassi TEXT;
        ''')

    # Adiciona a coluna 'numero_renavam' na tabela carros se ela não existir
    c.execute('''
        PRAGMA table_info(carros);
    ''')
    columns_carros = [col[1] for col in c.fetchall()]
    if 'numero_renavam' not in columns_carros:
        c.execute('''
            ALTER TABLE carros ADD COLUMN numero_renavam TEXT;
        ''')
    
    # Adiciona a coluna 'ano_veiculo' na tabela carros se ela não existir
    c.execute('''
        PRAGMA table_info(carros);
    ''')
    columns_carros = [col[1] for col in c.fetchall()]
    if 'ano_veiculo' not in columns_carros:
        c.execute('''
            ALTER TABLE carros ADD COLUMN ano_veiculo INTEGER;
        ''')

    # Adiciona a coluna 'km_troca_oleo' na tabela carros se ela não existir
    c.execute('''
        PRAGMA table_info(carros);
    ''')
    columns_carros = [col[1] for col in c.fetchall()]
    if 'km_troca_oleo' not in columns_carros:
        c.execute('''
            ALTER TABLE carros ADD COLUMN km_troca_oleo INTEGER DEFAULT 10000;
        ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            cpf TEXT UNIQUE,
            cnh TEXT,
            validade_cnh DATE,
            telefone TEXT,
            endereco TEXT,
            observacoes TEXT,
            status TEXT DEFAULT 'Ativo'
        )
    ''')

    # Adiciona a coluna 'status' na tabela clientes se ela não existir
    c.execute('''
        PRAGMA table_info(clientes);
    ''')
    columns = [col[1] for col in c.fetchall()]
    if 'status' not in columns:
        c.execute('''
            ALTER TABLE clientes ADD COLUMN status TEXT DEFAULT 'Ativo';
        ''')

    c.execute('''
        CREATE TABLE IF NOT EXISTS reservas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            carro_id INTEGER,
            cliente_id INTEGER,
            data_inicio DATE,
            data_fim DATE,
            -- NOVO CAMPO: Gerencia o fluxo de reserva (Reservada, Locada, Finalizada)
            reserva_status TEXT DEFAULT 'Reservada',
            status TEXT,
            custo_lavagem REAL DEFAULT 0,
            valor_total REAL DEFAULT 0,
            km_saida INTEGER,
            km_volta INTEGER,
            km_franquia INTEGER DEFAULT 300, -- Novo campo para franquia de KM
            adiantamento REAL DEFAULT 0.0, -- NOVO CAMPO: Valor de adiantamento
            valor_multas REAL DEFAULT 0.0, -- Novo campo para multas
            valor_danos REAL DEFAULT 0.0, -- Novo campo para danos ao veículo
            valor_outros REAL DEFAULT 0.0, -- Novo campo para outros custos extras
            FOREIGN KEY(carro_id) REFERENCES carros(id),
            FOREIGN KEY(cliente_id) REFERENCES clientes(id)
        )
    ''')

    # Adiciona a coluna 'km_franquia' na tabela reservas se ela não existir
    c.execute('''
        PRAGMA table_info(reservas);
    ''')
    columns_reservas = [col[1] for col in c.fetchall()]
    if 'km_franquia' not in columns_reservas:
        c.execute('''
            ALTER TABLE reservas ADD COLUMN km_franquia INTEGER DEFAULT 300;
        ''')
    
    # Adiciona a coluna 'adiantamento' na tabela reservas se ela não existir
    c.execute('''
        PRAGMA table_info(reservas);
    ''')
    columns_reservas_adiantamento = [col[1] for col in c.fetchall()]
    if 'adiantamento' not in columns_reservas_adiantamento:
        c.execute('''
            ALTER TABLE reservas ADD COLUMN adiantamento REAL DEFAULT 0.0;
        ''')
    
    # Adiciona a coluna 'valor_multas' na tabela reservas se ela não existir
    c.execute('''
        PRAGMA table_info(reservas);
    ''')
    columns_reservas_multas = [col[1] for col in c.fetchall()]
    if 'valor_multas' not in columns_reservas_multas:
        c.execute('''
            ALTER TABLE reservas ADD COLUMN valor_multas REAL DEFAULT 0.0;
        ''')

    # Adiciona a coluna 'valor_danos' na tabela reservas se ela não existir
    c.execute('''
        PRAGMA table_info(reservas);
    ''')
    columns_reservas_danos = [col[1] for col in c.fetchall()]
    if 'valor_danos' not in columns_reservas_danos:
        c.execute('''
            ALTER TABLE reservas ADD COLUMN valor_danos REAL DEFAULT 0.0;
        ''')

    # Adiciona a coluna 'valor_outros' na tabela reservas se ela não existir
    c.execute('''
        PRAGMA table_info(reservas);
    ''')
    columns_reservas_outros = [col[1] for col in c.fetchall()]
    if 'valor_outros' not in columns_reservas_outros:
        c.execute('''
            ALTER TABLE reservas ADD COLUMN valor_outros REAL DEFAULT 0.0;
        ''')
    
    conn.commit()
    conn.close()


def run_query(query, params=(), fetch=False):
    conn = sqlite3.connect('locadora_v2.db', detect_types=sqlite3.PARSE_DECLTYPES)
    c = conn.cursor()
    try:
        c.execute(query, params)
        if fetch:
            data = c.fetchall()
            headers = [description[0] for description in c.description]
            conn.close()
            return pd.DataFrame(data, columns=headers)
        conn.commit()
        if query.strip().upper().startswith("INSERT"):
            last_id = c.lastrowid
            conn.close()
            return last_id
        conn.close()
        return None
    except Exception as e:
        conn.close()
        return str(e)


def run_query_dataframe(query, params=()):
    """Executa uma query SELECT e retorna um DataFrame, ou um DataFrame vazio em caso de erro."""
    conn = sqlite3.connect('locadora_v2.db', detect_types=sqlite3.PARSE_DECLTYPES)
    c = conn.cursor()
    try:
        c.execute(query, params)
        data = c.fetchall()
        headers = [description[0] for description in c.description]
        conn.close()
        return pd.DataFrame(data, columns=headers)
    except Exception as e:
        st.error(f"Erro ao executar consulta: {e}")
        conn.close()
        return pd.DataFrame() # Retorna um DataFrame vazio em caso de erro


def gerar_recibo_para_download(reserva_id):
    # 1. Buscar dados da reserva
    query_reserva = """
        SELECT 
            r.id, r.data_inicio, r.data_fim, r.km_saida, r.km_volta, r.km_franquia,
            r.custo_lavagem, r.valor_multas, r.valor_danos, r.valor_outros, r.adiantamento, r.valor_total,
            cl.nome AS cliente_nome, cl.cpf AS cliente_cpf, cl.telefone AS cliente_telefone,
            c.modelo AS carro_modelo, c.placa AS carro_placa, c.cor AS carro_cor, c.preco_km AS carro_preco_km, c.diaria AS carro_diaria,
            c.numero_chassi AS carro_chassi, c.numero_renavam AS carro_renavam
        FROM reservas r
        JOIN clientes cl ON r.cliente_id = cl.id
        JOIN carros c ON r.carro_id = c.id
        WHERE r.id = ?
    """
    reserva_df = run_query_dataframe(query_reserva, (reserva_id,))

    if reserva_df.empty:
        st.error(f"Reserva com ID {reserva_id} não encontrada.")
        return None

    reserva_data = reserva_df.iloc[0].to_dict()

    # Preparar dados do cliente
    cliente = {
        'nome': reserva_data['cliente_nome'],
        'cpf': reserva_data['cliente_cpf'],
        'telefone': reserva_data['cliente_telefone']
    }

    # Preparar dados do carro
    carro = {
        'modelo': reserva_data['carro_modelo'],
        'placa': reserva_data['carro_placa'],
        'cor': reserva_data['carro_cor'],
        'preco_km': reserva_data['carro_preco_km'],
        'diaria': reserva_data['carro_diaria'],
        'chassi': reserva_data['carro_chassi'],
        'renavam': reserva_data['carro_renavam']
    }

    # Calcular dias de cobrança
    data_inicio_dt = pd.to_datetime(reserva_data['data_inicio']).date()
    data_fim_dt = pd.to_datetime(reserva_data['data_fim']).date()
    dias_cobranca = (data_fim_dt - data_inicio_dt).days

    # Preparar dados para o recibo
    recibo_dados = {
        'data_inicio': data_inicio_dt,
        'data_fim': data_fim_dt,
        'km_saida': reserva_data['km_saida'],
        'km_volta': reserva_data['km_volta'],
        'km_franquia': reserva_data['km_franquia'] if reserva_data['km_franquia'] is not None else 0.0,
        'dias_cobranca': dias_cobranca,
        'custo_diarias': reserva_data['carro_diaria'] * dias_cobranca,
        'custo_km': max(0, reserva_data['km_volta'] - reserva_data['km_saida'] - (reserva_data['km_franquia'] if reserva_data['km_franquia'] is not None else 0.0)) * reserva_data['carro_preco_km'],
        'valor_lavagem': reserva_data['custo_lavagem'] if reserva_data['custo_lavagem'] is not None else 0.0,
        'valor_multas': reserva_data['valor_multas'] if reserva_data['valor_multas'] is not None else 0.0,
        'valor_danos': reserva_data['valor_danos'] if reserva_data['valor_danos'] is not None else 0.0,
        'valor_outros': reserva_data['valor_outros'] if reserva_data['valor_outros'] is not None else 0.0,
        'adiantamento': reserva_data['adiantamento'] if reserva_data['adiantamento'] is not None else 0.0,
        'total_final': reserva_data['valor_total']
    }

    # Gerar o PDF
    try:
        pdf_bytes = gerar_recibo_pdf(cliente, carro, recibo_dados)
        return pdf_bytes
    except Exception as e:
        st.error(f"Erro ao gerar recibo PDF: {e}")
        return None

# --- INTERFACE ---
# Título da Página com acento
st.set_page_config(page_title="Locadora Iguacu Veiculos", layout="wide", page_icon="🚗")

# Informações do usuário logado
st.sidebar.markdown("---")
st.sidebar.markdown(f"**👤 {current_user['full_name']}**")
st.sidebar.markdown(f"**🔒 {current_user['role'].title()}**")

if st.sidebar.button("🚪 Logout", key="logout_btn"):
    logout()

st.sidebar.title("🚗 Menu Principal")

# Menu base
menu_options = ["Dashboard", "Clientes", "Frota (Carros)",
                "1. Reservar Veículo", "2. Entrega do Veículo",
                "Devolução", "Histórico", "Relatórios", "Backup"]

# Adicionar gerenciamento de usuários apenas para admin
if check_permission('manage_users'):
    menu_options.append("👥 Gerenciar Usuários")

menu = st.sidebar.radio("Ir para", menu_options, key="main_menu_selector")

# 1. DASHBOARD
if menu == "Dashboard":
    st.title("📊 Painel Gerencial e Agenda do Dia")

    # 1. Métricas Principais
    df_carros = run_query("SELECT * FROM carros", fetch=True)
    df_reservas = run_query("SELECT * FROM reservas", fetch=True)

    df_carros = df_carros if not isinstance(df_carros, str) else pd.DataFrame()
    df_reservas = df_reservas if not isinstance(df_reservas, str) else pd.DataFrame()

    # --- CÁLCULO DE FATURAMENTO MENSAL ---
    hoje = datetime.now()
    primeiro_dia_mes = hoje.replace(day=1).strftime('%Y-%m-%d')
    proximo_mes = hoje.replace(day=28) + timedelta(days=4)
    ultimo_dia_mes = (proximo_mes - timedelta(days=proximo_mes.day)).strftime('%Y-%m-%d')

    # A consulta de faturamento usa o 'status' original 'Finalizada'
    query_faturamento_mensal = f"""
        SELECT SUM(valor_total) FROM reservas 
        WHERE status = 'Finalizada' 
        AND data_fim BETWEEN '{primeiro_dia_mes}' AND '{ultimo_dia_mes}'
    """

    conn = sqlite3.connect('locadora_v2.db')
    c = conn.cursor()
    c.execute(query_faturamento_mensal)
    faturamento_mensal_resultado = c.fetchone()[0]
    conn.close()

    faturamento_mensal = faturamento_mensal_resultado if faturamento_mensal_resultado else 0.0
    # ---------------------------------------------

    hoje_str = date.today().strftime('%Y-%m-%d')

    # Busca Locados (reserva_status = 'Locada')
    query_locados = f"""
        SELECT 
            c.modelo, c.placa, r.data_fim, cl.nome as cliente 
        FROM reservas r
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Locada'
    """
    df_locados = run_query(query_locados, fetch=True)
    df_locados = df_locados if not isinstance(df_locados, str) else pd.DataFrame()

    # Busca Reservados (reserva_status = 'Reservada')
    query_reservados = f"""
        SELECT 
            c.modelo, c.placa, r.data_inicio, cl.nome as cliente 
        FROM reservas r
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Reservada'
    """
    df_reservados = run_query(query_reservados, fetch=True)
    df_reservados = df_reservados if not isinstance(df_reservados, str) else pd.DataFrame()

    # Consultas de Agenda (usando o novo reserva_status)
    # Devoluções Previstas são carros Locados (que saíram) e devem voltar hoje
    query_entradas_hoje = f"""
        SELECT c.modelo, c.placa, cl.nome AS cliente, r.data_fim
        FROM reservas r 
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Locada' AND date(r.data_fim) = date('{hoje_str}')
    """
    entradas_hoje = run_query(query_entradas_hoje, fetch=True)

    # Saídas Previstas são reservas que precisam ser entregues hoje
    query_saidas_hoje = f"""
        SELECT c.modelo, c.placa, cl.nome AS cliente, r.data_inicio
        FROM reservas r 
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.reserva_status = 'Reservada' AND date(r.data_inicio) = date('{hoje_str}')
    """
    saidas_hoje = run_query(query_saidas_hoje, fetch=True)

    df_entradas = entradas_hoje if not isinstance(entradas_hoje, str) else pd.DataFrame()
    df_saidas = saidas_hoje if not isinstance(saidas_hoje, str) else pd.DataFrame()

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Veículos na Frota", len(df_carros))
    col2.metric("Carros Locados Agora", len(df_locados))
    col3.metric("Carros Reservados", len(df_reservados))
    col4.metric(f"Faturamento {hoje.strftime('%b/%Y')}", formatar_moeda(faturamento_mensal))
    col5.metric("Devoluções Previstas Hoje", len(df_entradas))

    st.divider()

    # --- CHECAGEM RÁPIDA DE DISPONIBILIDADE (CORRIGIDO) ---
    st.subheader("🗓️ Verificação Rápida de Disponibilidade")

    col_data1, col_data2 = st.columns(2)

    data_inicio_check = col_data1.date_input("Início da Locação", date.today(), key="check_inicio")
    data_fim_check = col_data2.date_input("Fim da Locação", data_inicio_check + timedelta(days=1),
                                          min_value=data_inicio_check, key="check_fim")
    #data_fim_check = col_data2.date_input("Fim da Locação", date.today() + timedelta(days=7),
                                          #min_value=data_inicio_check, key="check_fim")

    if data_inicio_check <= data_fim_check:

        # CORREÇÃO CRÍTICA APLICADA AQUI: data_fim >= DATE('{inicio}', '-1 day')
        # Garante que, se a reserva termina no dia 27, a checagem só causa conflito até o dia 27 (28-1),
        # permitindo o aluguel no dia 28.

        query_check = f"""
            SELECT modelo, placa, diaria, preco_km 
            FROM carros 
            WHERE status NOT IN ('Indisponível', 'Excluído')  -- <<< Considera todos, exceto Indisponível
            AND id NOT IN (
                SELECT carro_id FROM reservas 
                WHERE reserva_status IN ('Reservada', 'Locada') -- <<< Apenas reservas ativas
                AND (data_inicio <= '{data_fim_check}' AND data_fim >= DATE('{data_inicio_check}', '+0 day'))
            )
        """
        livres_check = run_query(query_check, fetch=True)


        if not isinstance(livres_check, str) and not livres_check.empty:
            st.success(
                f"✅ {len(livres_check)} Veículos Disponíveis de {data_inicio_check.strftime('%d/%m')} a {data_fim_check.strftime('%d/%m')}.")

            # Formatando as moedas para exibição
            livres_check['Diária (R$)'] = livres_check['diaria'].apply(lambda x: formatar_moeda(x).replace('R$ ', ''))
            livres_check['Preço/KM (R$)'] = livres_check['preco_km'].apply(
                lambda x: formatar_moeda(x).replace('R$ ', ''))

            st.dataframe(
                livres_check.rename(columns={'modelo': 'Modelo', 'placa': 'Placa'})[
                    ['Modelo', 'Placa', 'Diária (R$)', 'Preço/KM (R$)']],
                width='stretch'
            )
        elif isinstance(livres_check, str):
            st.error(f"Erro ao consultar disponibilidade: {livres_check}")
        else:
            st.warning("⚠️ Nenhum carro disponível para o período selecionado.")

    st.divider()

    # 2. Situação de Carros Locados e Reservados
    col_status1, col_status2 = st.columns(2)

    with col_status1:
        st.subheader("Situação da Frota: Carros Locados")
        if not df_locados.empty:
            df_locados['data_fim'] = pd.to_datetime(df_locados['data_fim']).dt.strftime('%d/%m/%Y')

            st.dataframe(
                df_locados.rename(columns={
                    'modelo': 'Modelo',
                    'placa': 'Placa',
                    'cliente': 'Cliente',
                    'data_fim': 'Devolução Prevista'
                }),
                width='stretch'
            )
        else:
            st.success("🎉 Nenhum veículo locado no momento.")

    with col_status2:
        st.subheader("Situação da Frota: Carros Reservados (Aguardando Entrega)")
        if not df_reservados.empty:
            df_reservados['data_inicio'] = pd.to_datetime(df_reservados['data_inicio']).dt.strftime('%d/%m/%Y')

            st.dataframe(
                df_reservados.rename(columns={
                    'modelo': 'Modelo',
                    'placa': 'Placa',
                    'cliente': 'Cliente',
                    'data_inicio': 'Data Prevista da Entrega'
                }),
                width='stretch'
            )
        else:
            st.info("Nenhuma reserva pendente de entrega.")

    st.divider()

    # 3. Agenda do Dia
    st.subheader("Agenda: Entradas e Saídas do Dia (HOJE)")
    col_agenda1, col_agenda2 = st.columns(2)

    with col_agenda1:
        st.markdown("##### 📥 Devoluções Previstas (HOJE)")
        if not df_entradas.empty:
            st.dataframe(df_entradas.rename(columns={'modelo': 'Modelo', 'placa': 'Placa', 'cliente': 'Cliente'}),
                         width='stretch')
        else:
            st.info("Nenhuma devolução agendada.")

    with col_agenda2:
        st.markdown("##### 📤 Entregas Agendadas (HOJE)")
        if not df_saidas.empty:
            st.dataframe(df_saidas.rename(columns={'modelo': 'Modelo', 'placa': 'Placa', 'cliente': 'Cliente'}),
                         width='stretch')
        else:
            st.info("Nenhuma nova locação agendada.")


# 2. CLIENTES (COMPLETO: CADASTRO, EDIÇÃO, EXCLUSÃO E OBSERVAÇÕES)
elif menu == "Clientes":
    st.title("👥 Gestão de Clientes")

    # Banco já foi inicializado no início da aplicação

    tab1, tab2 = st.tabs(["Cadastrar Novo", "Ver / Editar Clientes"])

    with tab1:
        st.subheader("Cadastro de Cliente")
        with st.form("form_cliente"):
            c1, c2 = st.columns(2)
            nome = c1.text_input("Nome Completo")
            cpf = c2.text_input("CPF")

            c3, c4, c5 = st.columns(3)
            cnh = c3.text_input("Número da CNH")
            # Define o valor padrão para evitar erro de inicialização se não houver data
            validade_cnh = c4.date_input("Validade CNH", min_value=date.today(), value=date.today())
            telefone = c5.text_input("Telefone")

            endereco = st.text_area("Endereço Completo")

            # --- NOVO CAMPO DE OBSERVAÇÕES (Não obrigatório) ---
            observacoes = st.text_area("Observações sobre o Cliente (Não obrigatório)")
            # ----------------------------------------------------

            submit = st.form_submit_button("Salvar Cliente", type="primary")

            if submit:
                # Validação de campos obrigatórios
                if not nome or not cpf or not cnh or not telefone:
                    st.error("⚠️ Os campos Nome, CPF, CNH e Telefone são obrigatórios!")
                else:
                    # Incluindo 'validade_cnh' e 'observacoes' na query INSERT
                    res = run_query(
                        "INSERT INTO clientes (nome, cpf, cnh, validade_cnh, telefone, endereco, observacoes) VALUES (?,?,?,?,?,?,?)",
                        (nome, cpf, cnh, validade_cnh, telefone, endereco, observacoes)
                    )
                    if isinstance(res, str):
                        st.error(f"Erro ao cadastrar (Verifique se o CPF já existe). Detalhe: {res}")
                    else:
                        st.toast(f"Cliente {nome} cadastrado com sucesso!", icon="✅")
                        st.success(f"Cliente {nome} cadastrado com sucesso!")
                        st.rerun()

    with tab2:
        # Filtra para não exibir clientes com status 'Removido' na lista principal
        df_clientes = run_query_dataframe(f"SELECT * FROM clientes WHERE status != '{STATUS_CLIENTE['REMOVIDO']}'")
        if not df_clientes.empty:

            # Formatação para exibição
            df_clientes_display = df_clientes.copy()
            df_clientes_display['validade_cnh'] = pd.to_datetime(df_clientes_display['validade_cnh']).dt.strftime(
                '%d/%m/%Y')

            st.dataframe(df_clientes_display, width='stretch')

            cliente_opcoes = df_clientes['id'].astype(str) + " - " + df_clientes['nome']
            opcoes_com_placeholder = ["Selecione o cliente..."] + cliente_opcoes.tolist()

            cliente_sel = st.selectbox("Selecione para Edição ou Exclusão", opcoes_com_placeholder)

            if cliente_sel != "Selecione o cliente...":
                id_cliente_sel = int(cliente_sel.split(" - ")[0])
                dados_atuais = df_clientes[df_clientes['id'] == id_cliente_sel].iloc[0]

                # --- FORMULÁRIO DE EDIÇÃO ---
                st.markdown("---")
                st.subheader(f"✏️ Editando: {dados_atuais['nome']}")

                # Prepara o valor da data da CNH
                if dados_atuais['validade_cnh']:
                    validade_cnh_atual = pd.to_datetime(dados_atuais['validade_cnh']).date()
                else:
                    validade_cnh_atual = date.today()

                with st.form("form_edit_cliente"):
                    # Colunas para campos que não devem ser alterados (CPF) ou são chave (CNH)
                    e1, e2 = st.columns(2)
                    up_nome = e1.text_input("Nome Completo", value=dados_atuais['nome'])
                    e2.text_input("CPF (Não Editável)", value=dados_atuais['cpf'], disabled=True)

                    e3, e4, e5 = st.columns(3)
                    up_cnh = e3.text_input("Número da CNH", value=dados_atuais['cnh'])
                    up_validade_cnh = e4.date_input("Validade CNH", value=validade_cnh_atual, min_value=date.today())
                    up_telefone = e5.text_input("Telefone", value=dados_atuais['telefone'])

                    up_endereco = st.text_area("Endereço Completo", value=dados_atuais['endereco'])

                    # Carrega o valor atual para observações (usando .fillna('') para evitar NaN)
                    up_observacoes = st.text_area("Observações", value=dados_atuais['observacoes'] if pd.notna(
                        dados_atuais['observacoes']) else '')

                    col_botoes = st.columns(2)

                    if col_botoes[0].form_submit_button("🔄 Atualizar Dados do Cliente", type="primary"):
                        if not up_nome or not up_cnh or not up_telefone:
                            st.error("⚠️ Os campos Nome, CNH e Telefone não podem ficar vazios.")
                        else:
                            run_query("""
                                UPDATE clientes 
                                SET nome=?, cnh=?, validade_cnh=?, telefone=?, endereco=?, observacoes=? 
                                WHERE id=?
                            """, (
                            up_nome, up_cnh, up_validade_cnh, up_telefone, up_endereco, up_observacoes, id_cliente_sel))
                            st.toast("Cliente atualizado!", icon="✔️")
                            st.success(f"Cliente **{up_nome}** atualizado com sucesso!")
                            st.rerun()

                    if col_botoes[1].form_submit_button("🗑️ Marcar como REMOVIDO"):
                        # --- CHECAGEM CRÍTICA DE RESERVAS ATIVAS (REPETIDA DA LÓGICA ANTERIOR) ---
                        reservas_ativas_check = run_query(
                            "SELECT COUNT(*) FROM reservas WHERE cliente_id=? AND reserva_status IN ('Reservada', 'Locada')",
                            (id_cliente_sel,),
                            fetch=True
                        ).iloc[0, 0]

                        if reservas_ativas_check > 0:
                            st.error(
                                f"❌ Não é possível remover. O cliente possui {reservas_ativas_check} reserva(s) Ativa(s). Finalize a devolução primeiro.")
                        else:
                            # Se não há reservas ativas, marca o cliente como 'Removido'
                            run_query("UPDATE clientes SET status=? WHERE id=?", (STATUS_CLIENTE['REMOVIDO'], id_cliente_sel))
                            st.toast("Cliente marcado como Removido!", icon="🗑️")
                            st.warning("Cliente marcado como **REMOVIDO** (Registro mantido para histórico).")
                            st.rerun()
        else:
            st.info("Nenhum cliente cadastrado ou ativo.")


# 3. FROTA (CARROS) - COMPLETO COM STATUS 'EXCLUÍDO'
elif menu == "Frota (Carros)":
    st.title("🚙 Gestão da Frota")

    tab1, tab2 = st.tabs(["Cadastrar Veículo", "Ver / Editar / Status"])

    with tab1:
        st.subheader("Cadastro de Novo Veículo")
        with st.form("cadastro_carro", clear_on_submit=False):
            col_a, col_b = st.columns(2)
            modelo = col_a.text_input("Modelo (ex: Fiat Mobi)")
            placa = col_b.text_input("Placa").upper()

            col_c, col_d = st.columns(2)
            cor = col_c.text_input("Cor")
            km = col_d.number_input("KM Atual", 0)

            col_e, col_f, col_g = st.columns(3)
            numero_chassi = col_e.text_input("Número do Chassi")
            numero_renavam = col_f.text_input("Número do Renavam")
            km_troca_oleo = col_g.number_input("KM da Próxima Troca de Óleo", min_value=0, value=10000)

            ano_veiculo = st.number_input("Ano do Veículo", min_value=1900, max_value=date.today().year + 1, value=date.today().year)

            col_h, col_i = st.columns(2)
            diaria = col_h.number_input("Valor Diária (R$)", 0.0)
            p_km = col_i.number_input("Custo por KM (R$)", 0.0)

            # Define o status inicial no cadastro
            status_inicial = st.selectbox("Status Inicial (Padrão)", options=list(STATUS_CARRO.values()), index=0)

            if st.form_submit_button("Salvar Carro", type="primary"):
                if not modelo or not placa or not cor or diaria <= 0 or not numero_chassi or not numero_renavam or not ano_veiculo:
                    st.error("⚠️ Preencha Modelo, Placa, Cor, Número do Chassi, Número do Renavam, Ano do Veículo e certifique-se que a Diária seja maior que zero.")
                else:
                    res = run_query(
                        "INSERT INTO carros (modelo, placa, cor, km_atual, diaria, preco_km, status, numero_chassi, numero_renavam, ano_veiculo, km_troca_oleo) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                        (modelo, placa, cor, km, diaria, p_km, status_inicial, numero_chassi, numero_renavam, ano_veiculo, km_troca_oleo)
                    )
                    if isinstance(res, str):
                        st.error(f"Erro ao cadastrar. Detalhe: {res}")
                    else:
                        st.toast("Veículo Cadastrado!", icon="✅")
                        st.success("Veículo Cadastrado!")
                        st.rerun()

    with tab2:
        df = run_query(f"SELECT * FROM carros WHERE status != '{STATUS_CARRO['EXCLUIDO']}'", fetch=True)
        if not isinstance(df, str) and not df.empty:
            st.subheader("Frota Atual")
            # Adiciona coluna calculada para KM até próxima troca de óleo
            df_display = df.copy()
            df_display['km_ate_proxima_troca'] = df_display.apply(
                lambda row: max(0, int(row['km_troca_oleo']) - int(row['km_atual'])) if pd.notna(row['km_troca_oleo']) and pd.notna(row['km_atual']) else 0,
                axis=1
            )

            # Exibe apenas as colunas principais incluindo a nova coluna calculada
            st.dataframe(df_display[['id', 'modelo', 'placa', 'cor', 'km_atual', 'km_troca_oleo', 'km_ate_proxima_troca', 'diaria', 'status', 'numero_chassi', 'numero_renavam', 'ano_veiculo']], width='stretch')

            carro_opcoes = df['id'].astype(str) + " - " + df['modelo'] + " (" + df['placa'] + ")"
            opcoes_com_placeholder = ["Selecione o veículo..."] + carro_opcoes.tolist()

            carro_sel = st.selectbox("Selecione Veículo para Ação", opcoes_com_placeholder)

            if carro_sel != "Selecione o veículo...":
                id_edit = int(carro_sel.split(" - ")[0])
                dados_atuais = df[df['id'] == id_edit].iloc[0]

                # --- FORMULÁRIO DE EDIÇÃO DE DADOS E STATUS ---
                st.markdown("---")
                st.subheader(f"✏️ Editando: {dados_atuais['modelo']} ({dados_atuais['placa']})")

                with st.form("form_edit_carro"):
                    # 1. Campos de dados
                    st.markdown("##### Dados do Veículo")
                    col_info1, col_info2, col_info3 = st.columns(3)
                    col_info1.text_input("Modelo", value=dados_atuais['modelo'], disabled=True)
                    col_info2.text_input("Placa", value=dados_atuais['placa'], disabled=True)
                    up_cor = col_info3.text_input("Cor", value=dados_atuais['cor'], key="up_cor")

                    col_chassi, col_renavam, col_ano = st.columns(3)
                    up_numero_chassi = col_chassi.text_input("Número do Chassi", value=dados_atuais['numero_chassi'], key="up_numero_chassi")
                    up_numero_renavam = col_renavam.text_input("Número do Renavam", value=dados_atuais['numero_renavam'], key="up_numero_renavam")
                    up_km_troca_oleo = col_ano.number_input("KM da Próxima Troca de Óleo", min_value=0, value=int(dados_atuais['km_troca_oleo']) if dados_atuais['km_troca_oleo'] is not None and not pd.isna(dados_atuais['km_troca_oleo']) else 10000, key="up_km_troca_oleo")

                    up_ano_veiculo = st.number_input("Ano do Veículo", min_value=1900, max_value=date.today().year + 1, value=int(dados_atuais['ano_veiculo']) if dados_atuais['ano_veiculo'] is not None and not pd.isna(dados_atuais['ano_veiculo']) else date.today().year, key="up_ano_veiculo")

                    st.markdown("##### Valores e Quilometragem")
                    col_a, col_b = st.columns(2)
                    up_diaria = col_a.number_input("Valor Diária (R$)", value=float(dados_atuais['diaria']),
                                                   key="up_diaria")
                    up_p_km = col_b.number_input("Custo por KM (R$)", value=float(dados_atuais['preco_km']),
                                                 key="up_p_km")
                    up_km = st.number_input("KM Atual", value=int(dados_atuais['km_atual']), key="up_km")

                    st.markdown("---")

                    # 2. Edição de Status
                    st.markdown("##### Status do Veículo")

                    # Encontra o índice do status atual
                    status_index = list(STATUS_CARRO.values()).index(dados_atuais['status']) if dados_atuais[
                                                                                                    'status'] in STATUS_CARRO.values() else 0

                    up_status = st.selectbox(
                        "Alterar Status:",
                        options=list(STATUS_CARRO.values()),
                        index=status_index,
                        key="up_status",
                        help="Define o estado do veículo."
                    )

                    col_botoes = st.columns(2)

                    if col_botoes[0].form_submit_button("🔄 Atualizar Dados e Status", type="primary"):
                        # Verificação de KM para prevenir regressão
                        if up_km < int(dados_atuais['km_atual']):
                            st.error("❌ O KM atualizado não pode ser menor que o KM registrado anteriormente.")
                        else:
                            # Atualiza todos os campos
                            run_query("UPDATE carros SET cor=?, diaria=?, preco_km=?, km_atual=?, status=?, numero_chassi=?, numero_renavam=?, ano_veiculo=?, km_troca_oleo=? WHERE id=?",
                                      (up_cor, up_diaria, up_p_km, up_km, up_status, up_numero_chassi, up_numero_renavam, up_ano_veiculo, up_km_troca_oleo, id_edit))
                            st.toast("Dados do veículo atualizados!", icon="✔️")
                            st.success(f"Veículo **{dados_atuais['modelo']}** atualizado para status **{up_status}**!")
                            st.rerun()

                    # Botão para retirar definitivamente da frota, mudando o status para EXCLUÍDO
                    if col_botoes[1].form_submit_button("🔥 Marcar como EXCLUÍDO (Retirada Definitiva)"):
                        # Verifica se o carro está locado ou reservado
                        if dados_atuais['status'] in [STATUS_CARRO['LOCADO'], STATUS_CARRO['RESERVADO']]:
                            st.error(
                                f"❌ Não é possível excluir. O carro está **{dados_atuais['status']}**. Finalize a pendência primeiro.")
                        else:
                            # Se não está em locação, define o status para 'EXCLUIDO'
                            run_query("UPDATE carros SET status=? WHERE id=?",
                                      (STATUS_CARRO['EXCLUIDO'], id_edit))
                            st.toast("Carro marcado como Excluído!", icon="🔥")
                            st.error("Veículo marcado como **EXCLUÍDO** (Registro mantido para histórico).")
                            st.rerun()
        else:
            st.info("Nenhum carro cadastrado na frota.")

# 4. 1. RESERVAR VEÍCULO (Apenas bloqueia o carro por data)
elif menu == "1. Reservar Veículo":
    st.title("📝 Gerenciamento de Reservas")

    tab_reservar, tab_gerenciar = st.tabs(["Reservar Novo Veículo", "Visualizar/Editar/Excluir Reservas"])

    with tab_reservar:
        st.subheader("Nova Reserva (Bloqueio de Data)")

        c1, c2 = st.columns(2)
        inicio = c1.date_input("Data Retirada (Previsão)", min_value=date.today(), key="reserva_inicio_novo")
        fim = c2.date_input("Data Devolução (Previsão)",inicio + timedelta(days=1), min_value=inicio,key="reserva_fim_novo")

        clientes_db = run_query_dataframe("SELECT * FROM clientes WHERE status = 'Ativo'") # Apenas clientes ativos

        if clientes_db.empty:
            st.warning("⚠️ Você precisa cadastrar clientes ativos na aba 'Clientes' antes de fazer uma reserva.")
        else:
            lista_clientes = clientes_db['id'].astype(str) + " - " + clientes_db['nome'] + " (CPF: " + clientes_db[
                'cpf'] + ")"
            opcoes_cliente_placeholder = ["Selecione o cliente..."] + lista_clientes.tolist()

            cliente_escolhido_str = st.selectbox("Selecione o Cliente", opcoes_cliente_placeholder, key="reserva_cliente_novo")

            if cliente_escolhido_str != "Selecione o cliente...":

                try:
                    cliente_id = int(cliente_escolhido_str.split(" - ")[0])
                    dados_cliente = clientes_db[clientes_db['id'] == cliente_id].iloc[0]
                except:
                    st.warning("Erro ao processar ID do cliente. Selecione novamente.")
                    cliente_id = None
                    dados_cliente = None

                if cliente_id and inicio <= fim:

                    # Consulta para carros DISPONÍVEIS no período (status 'Disponível' E sem conflito)
                    query = f"""
                        SELECT * FROM carros 
                        WHERE status NOT IN ('{STATUS_CARRO['INDISPONIVEL']}', '{STATUS_CARRO['EXCLUIDO']}') 
                        AND id NOT IN (
                            SELECT carro_id FROM reservas 
                            WHERE reserva_status IN ('Reservada', 'Locada') 
                            AND (data_inicio <= '{fim}' AND data_fim >= DATE('{inicio}', '+0 day'))
                        )
                    """
                    livres = run_query_dataframe(query)

                    if not livres.empty:
                        st.info(f"{len(livres)} veículos disponíveis para o período.")

                        carro_opcoes = livres['id'].astype(str) + " - " + livres['modelo'] + " - " + livres['placa']
                        opcoes_carro_placeholder = ["Selecione o veículo..."] + carro_opcoes.tolist()

                        carro_escolhido_str = st.selectbox("Escolha o Veículo", opcoes_carro_placeholder, key="reserva_carro_novo")

                        if carro_escolhido_str != "Selecione o veículo...":

                            try:
                                carro_id = int(carro_escolhido_str.split(" - ")[0])
                                dados_carro = livres[livres['id'] == carro_id].iloc[0]
                            except:
                                st.warning("Erro ao processar ID do veículo. Selecione novamente.")
                                carro_id = None
                                dados_carro = None

                            if carro_id:
                                dias = (fim - inicio).days
                                valor_previsto = dados_carro['diaria'] * max(dias, 1)  # Mínimo 1 dia

                                st.metric("Valor Previsto (Diárias)", formatar_moeda(valor_previsto),
                                          f"{max(dias, 1)} dias")
                                
                                # NOVO CAMPO: Franquia de KM
                                km_franqui_input = st.number_input("KM de Franquia (primeiros KM's gratuitos)", min_value=0, value=300,
                                                                    help="Defina a quantidade de KM que serão gratuitos para esta locação.", key="reserva_km_franquia_novo")
                                
                                # NOVO CAMPO: Adiantamento da locação (50% do valor previsto como padrão)
                                adiantamento_default = valor_previsto * 0.5
                                adiantamento_input = st.number_input("Valor de Adiantamento (R$)", min_value=0.0, 
                                                                     value=min(adiantamento_default, valor_previsto), # Garante que o adiantamento não seja maior que o valor previsto
                                                                     step=10.0, format="%.2f", key="reserva_adiantamento_novo")

                                if st.button("✅ Confirmar Reserva (Apenas Bloqueio de Data)", type="primary", key="confirmar_reserva_novo"):

                                    # Salvamos o KM ATUAL do carro no campo km_saida da reserva
                                    km_saida_salvar = int(dados_carro['km_atual'])

                                    # INCLUSÃO: Adicionando o campo reserva_status='Reservada', km_franquia e adiantamento
                                    reserv_id = run_query(
                                        "INSERT INTO reservas (carro_id, cliente_id, data_inicio, data_fim, status, reserva_status, km_saida, km_franquia, adiantamento, valor_multas, valor_danos, valor_outros) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                                        (carro_id, cliente_id, inicio, fim, 'Ativa', 'Reservada', km_saida_salvar, km_franqui_input, adiantamento_input, 0.0, 0.0, 0.0)
                                    )

                                    if isinstance(reserv_id, int):
                                        # ATENÇÃO: NÃO ATUALIZAMOS o status do carro aqui. Ele permanece 'Disponível'.
                                        st.toast("Reserva Confirmada! Aguardando Entrega.", icon="🎉")
                                        st.success(
                                            f"Reserva #{reserv_id} Confirmada. Prossiga para '2. Entrega do Veículo' para finalizar a locação.")
                                        st.rerun()
                                    else:
                                        st.error(f"Erro ao salvar a reserva: {reserv_id}")

                    elif isinstance(livres, str):
                        st.error(f"Erro na consulta de carros disponíveis: {livres}")
                    else:
                        st.warning("Sem carros disponíveis para estas datas no período.")

    with tab_gerenciar:
        st.subheader("Visualizar, Editar ou Excluir Reservas")

        # Buscar todas as reservas que não estão 'Finalizada' ou 'Cancelada'
        reservas_gerenciar = run_query_dataframe("""
            SELECT 
                r.id, cl.nome AS Cliente, c.modelo AS Veiculo, c.placa AS Placa, 
                r.data_inicio, r.data_fim, r.reserva_status, r.km_franquia, r.adiantamento, 
                r.carro_id, r.cliente_id, c.diaria, c.preco_km, c.km_atual, r.valor_multas, r.valor_danos, r.valor_outros
            FROM reservas r 
            JOIN clientes cl ON r.cliente_id = cl.id 
            JOIN carros c ON r.carro_id = c.id
            WHERE r.reserva_status IN ('Reservada', 'Locada')
            ORDER BY r.data_inicio ASC
        """)

        if reservas_gerenciar.empty:
            st.info("Nenhuma reserva ativa para gerenciar.")
        else:
            # Formatar para exibição
            reservas_gerenciar_display = reservas_gerenciar.copy()
            reservas_gerenciar_display['data_inicio'] = pd.to_datetime(
                reservas_gerenciar_display['data_inicio']).dt.strftime('%d/%m/%Y')
            reservas_gerenciar_display['data_fim'] = pd.to_datetime(reservas_gerenciar_display['data_fim']).dt.strftime(
                '%d/%m/%Y')
            reservas_gerenciar_display['adiantamento'] = reservas_gerenciar_display['adiantamento'].apply(
                formatar_moeda)
            reservas_gerenciar_display['valor_multas'] = reservas_gerenciar_display['valor_multas'].apply(
                formatar_moeda)
            reservas_gerenciar_display['valor_danos'] = reservas_gerenciar_display['valor_danos'].apply(formatar_moeda)
            reservas_gerenciar_display['valor_outros'] = reservas_gerenciar_display['valor_outros'].apply(
                formatar_moeda)
            reservas_gerenciar_display.rename(columns={
                'reserva_status': 'Status da Reserva',
                'km_franquia': 'KM Franquia',
                'adiantamento': 'Adiantamento',
                'valor_multas': 'Multas',
                'valor_danos': 'Danos',
                'valor_outros': 'Outros Custos'
            }, inplace=True)

            st.dataframe(
                reservas_gerenciar_display[
                    ['Cliente', 'Veiculo', 'Placa', 'data_inicio', 'data_fim', 'Status da Reserva', 'KM Franquia',
                     'Adiantamento', 'Multas', 'Danos', 'Outros Custos']],
                width="stretch"
            )

            # Seleção de reserva para edição/exclusão
            opcoes_reserva = reservas_gerenciar.apply(
                lambda
                    x: f"ID {x['id']} - {x['Cliente']} ({x['Veiculo']} - {x['Placa']}) | De {pd.to_datetime(x['data_inicio']).strftime('%d/%m')} a {pd.to_datetime(x['data_fim']).strftime('%d/%m')}",
                axis=1
            )
            opcoes_com_placeholder_reserva = ["Selecione uma reserva para gerenciar..."] + opcoes_reserva.tolist()

            reserva_selecionada_str = st.selectbox("Selecionar Reserva", opcoes_com_placeholder_reserva,
                                                   key="gerenciar_reserva_sel")

            if reserva_selecionada_str != "Selecione uma reserva para gerenciar...":
                id_reserva_gerenciar = int(reserva_selecionada_str.split(" - ")[0].replace("ID ", ""))
                reserva_atual = reservas_gerenciar[reservas_gerenciar['id'] == id_reserva_gerenciar].iloc[0].to_dict()

                st.markdown("---")
                st.subheader(f"✏️ Editando Reserva ID: {reserva_atual['id']}")

                with st.form(key=f"form_edit_reserva_{reserva_atual['id']}"):
                    # Cliente e Veículo atuais (somente leitura)
                    st.text_input("Cliente", value=reserva_atual['Cliente'], disabled=True)
                    st.text_input("Veículo Atual", value=f"{reserva_atual['Veiculo']} ({reserva_atual['Placa']})",
                                  disabled=True)
                    st.text_input("Status da Reserva", value=reserva_atual['reserva_status'], disabled=True)

                    col_edit_data1, col_edit_data2 = st.columns(2)
                    up_data_inicio = col_edit_data1.date_input("Nova Data de Retirada", value=pd.to_datetime(
                        reserva_atual['data_inicio']).date(), key=f"up_data_inicio_{reserva_atual['id']}")
                    up_data_fim = col_edit_data2.date_input("Nova Data de Devolução",
                                                            value=pd.to_datetime(reserva_atual['data_fim']).date(),
                                                            min_value=up_data_inicio,
                                                            key=f"up_data_fim_{reserva_atual['id']}")

                    up_km_franquia = st.number_input("Nova KM de Franquia", min_value=0,
                                                     value=reserva_atual['km_franquia'],
                                                     key=f"up_km_franquia_{reserva_atual['id']}")
                    up_adiantamento = st.number_input("Novo Valor de Adiantamento (R$)", min_value=0.0,
                                                      value=reserva_atual['adiantamento'], step=10.0, format="%.2f",
                                                      key=f"up_adiantamento_{reserva_atual['id']}")

                    st.markdown("##### Trocar Veículo (Opcional)")
                    # Consulta para carros DISPONÍVEIS no período (excluindo o carro atual da reserva)
                    query_carros_disponiveis = f"""
                        SELECT id, modelo, placa FROM carros 
                        WHERE status NOT IN ('{STATUS_CARRO['INDISPONIVEL']}', '{STATUS_CARRO['EXCLUIDO']}') 
                        AND id != {reserva_atual['carro_id']}
                        AND id NOT IN (
                            SELECT carro_id FROM reservas 
                            WHERE reserva_status IN ('Reservada', 'Locada') 
                            AND (data_inicio <= '{up_data_fim}' AND data_fim >= DATE('{up_data_inicio}', '+0 day'))
                        )
                    """
                    carros_disponiveis = run_query_dataframe(query_carros_disponiveis)

                    if carros_disponiveis.empty:
                        st.warning("Nenhum outro veículo disponível para troca neste período.")
                        lista_carros_troca = [
                            f"Manter Veículo Atual ({reserva_atual['Veiculo']} - {reserva_atual['Placa']})"]
                    else:
                        opcoes_troca = carros_disponiveis.apply(lambda x: f"{x['id']} - {x['modelo']} ({x['placa']})",
                                                                axis=1).tolist()
                        lista_carros_troca = [
                                                 f"Manter Veículo Atual ({reserva_atual['Veiculo']} - {reserva_atual['Placa']})"] + opcoes_troca

                    veiculo_troca_str = st.selectbox("Selecionar Novo Veículo (se desejar trocar)", lista_carros_troca,
                                                     key=f"veiculo_troca_sel_{reserva_atual['id']}")

                    # Botões de ação
                    col_edit_botoes = st.columns(3)
                    update_reserva = col_edit_botoes[0].form_submit_button("🔄 Atualizar Reserva", type="primary",
                                                                           key=f"btn_update_reserva_{reserva_atual['id']}")
                    cancel_reserva = col_edit_botoes[1].form_submit_button("🗑️ Cancelar Reserva",
                                                                           key=f"btn_cancel_reserva_{reserva_atual['id']}")
                    gerar_contrato = col_edit_botoes[2].form_submit_button("📄 Gerar Contrato",
                                                                           key=f"btn_gerar_contrato_{reserva_atual['id']}")

                    if update_reserva:
                        novo_carro_id = reserva_atual['carro_id']
                        if veiculo_troca_str != f"Manter Veículo Atual ({reserva_atual['Veiculo']} - {reserva_atual['Placa']})":
                            novo_carro_id = int(veiculo_troca_str.split(" - ")[0])

                        run_query("""
                            UPDATE reservas 
                            SET data_inicio=?, data_fim=?, km_franquia=?, adiantamento=?, carro_id=?, valor_multas=?, valor_danos=?, valor_outros=? 
                            WHERE id=?
                        """, (
                            up_data_inicio,
                            up_data_fim,
                            up_km_franquia,
                            up_adiantamento,
                            novo_carro_id,
                            reserva_atual['valor_multas'],
                            reserva_atual['valor_danos'],
                            reserva_atual['valor_outros'],
                            reserva_atual['id']
                        ))
                        st.toast("Reserva atualizada com sucesso!", icon="✔️")
                        st.success(f"Reserva ID **{reserva_atual['id']}** atualizada.")
                        st.rerun()

                    if cancel_reserva:
                        # Verifica se a reserva está locada (carro com o cliente)
                        if reserva_atual['reserva_status'] == 'Locada':
                            st.error(
                                "❌ Não é possível cancelar uma reserva que já está Locada. Realize a Devolução primeiro.")
                        else:
                            # Se a reserva não estiver locada, muda o status para 'Cancelada'
                            run_query("UPDATE reservas SET reserva_status='Cancelada' WHERE id=?",
                                      (reserva_atual['id'],))
                            st.toast("Reserva cancelada!", icon="🗑️")
                            st.warning(f"Reserva ID **{reserva_atual['id']}** cancelada.")
                            st.rerun()

                    if gerar_contrato:
                        # Verificar se a reserva já foi entregue (veículo já foi locado)
                        if reserva_atual['reserva_status'] != 'Locada':
                            st.error("❌ O contrato só pode ser gerado após a entrega efetiva do veículo (status 'Locada'). Use o menu '2. Entrega do Veículo' para entregar o veículo primeiro.")
                        else:
                            # Buscar dados completos do cliente e do carro para gerar o contrato
                            dados_cliente_contrato = run_query_dataframe("SELECT * FROM clientes WHERE id=?",
                                                                         (reserva_atual['cliente_id'],))
                            dados_carro_contrato = run_query_dataframe("SELECT * FROM carros WHERE id=?",
                                                                       (reserva_atual['carro_id'],))

                            if not dados_cliente_contrato.empty and not dados_carro_contrato.empty:
                                cliente_dict = dados_cliente_contrato.iloc[0].to_dict()
                                carro_dict = dados_carro_contrato.iloc[0].to_dict()
                                # Adiciona os campos necessários para o PDF
                                carro_dict['chassi'] = carro_dict.get('numero_chassi')
                                carro_dict['renavam'] = carro_dict.get('numero_renavam')

                                # IMPORTANTE: Usar a data real da entrega (data_inicio da reserva quando foi locada)
                                # e não a data da reserva original
                                data_entrega_real = pd.to_datetime(reserva_atual['data_inicio']).date()
                                data_devolucao_prevista = pd.to_datetime(reserva_atual['data_fim']).date()

                                # Gerar o PDF do contrato
                                pdf_bytes_contrato = gerar_contrato_pdf(
                                    cliente_dict,
                                    carro_dict,
                                    data_entrega_real,  # Data real da entrega do veículo
                                    data_devolucao_prevista  # Data prevista de devolução
                                )

                                # Salvar no Session State
                                st.session_state.pdf_para_download = pdf_bytes_contrato
                                st.session_state.pdf_file_name = f"contrato_{cliente_dict['nome']}_{carro_dict['placa']}.pdf"
                                st.toast("Contrato gerado! Clique no botão de download abaixo.", icon="📄")
                                st.success("✅ Contrato gerado com sucesso! Use o botão de download abaixo.")
                                st.rerun()
                            else:
                                st.error("Erro ao buscar dados para gerar o contrato.")

        # Botão de download do contrato (fora do formulário, para aparecer após gerar)
        if 'pdf_para_download' in st.session_state and st.session_state.pdf_para_download:
        #if st.session_state.pdf_para_download:
            st.markdown("---")
            st.download_button(
                label="📥 Baixar Contrato em PDF",
                data=st.session_state.pdf_para_download,
                file_name=st.session_state.pdf_file_name,
                mime="application/pdf",
                key="download_contrato_edicao_reserva"
            )
            # Limpa o estado após exibir o botão
            if st.button("✅ Concluído - Limpar", key="limpar_download_contrato"):
                st.session_state.pdf_para_download = None
                st.session_state.pdf_file_name = None
                st.rerun()

# 5. 2. ENTREGA DO VEÍCULO (Confirma KM, Altera Status e Imprime Contrato)
elif menu == "2. Entrega do Veículo":
    st.title("🔑 2. Confirmação de Entrega (Início da Locação)")
    st.info("Selecione uma reserva ativa e confirme a quilometragem para iniciar a locação.")

    # Inicializa o Session State para o PDF, se necessário
    if 'pdf_para_download' not in st.session_state:
        st.session_state.pdf_para_download = None
        st.session_state.pdf_file_name = None

    # MUDANÇA NO FILTRO: Busca reservas ATIVAS que estão com reserva_status='Reservada'
    query_entregar = """
        SELECT 
            r.id, cl.nome, c.modelo, c.placa, c.km_atual, r.carro_id, r.cliente_id, 
            r.data_inicio, r.data_fim  
        FROM reservas r 
        JOIN carros c ON r.carro_id = c.id 
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status='Ativa' AND r.reserva_status='Reservada'
        ORDER BY r.data_inicio ASC
    """
    reservas_entregar = run_query(query_entregar, fetch=True)

    if isinstance(reservas_entregar, str):
        st.error(f"Erro no banco de dados: {reservas_entregar}")
    elif not reservas_entregar.empty:
        # CORREÇÃO DA FORMATAÇÃO E INCLUSÃO DA DATA
        opcoes = reservas_entregar.apply(
            lambda
                x: f"ID {x['id']} - {x['nome']} ({x['modelo']} - {x['placa']}) | Retirada Prevista: {pd.to_datetime(x['data_inicio']).strftime('%d/%m/%Y')}",
            axis=1
        )
        opcoes_com_placeholder = ["Selecione a reserva para entrega..."] + opcoes.tolist()

        sel = st.selectbox("Reserva Pendente de Entrega", opcoes_com_placeholder)

        if sel != "Selecione a reserva para entrega...":

            try:
                # O ID da reserva vem formatado como "ID [ID] - ..."
                id_reserva_sel = int(sel.split(" - ")[0].replace("ID ", ""))
                reserva = reservas_entregar[reservas_entregar['id'] == id_reserva_sel].iloc[0]
            except Exception as e:
                st.warning(f"Erro ao processar ID da reserva. Selecione novamente. ({e})")
                reserva = None

            if reserva is not None:
                carro_id = int(reserva['carro_id'])
                cliente_id = int(reserva['cliente_id'])

                # Buscar dados completos para o contrato
                dados_carro = run_query("SELECT * FROM carros WHERE id=?", (carro_id,), fetch=True).iloc[0].to_dict()
                # Garante que os campos chassi, renavam, cor e ano_veiculo existem, mesmo que None
                dados_carro['chassi'] = dados_carro.get('numero_chassi')
                dados_carro['renavam'] = dados_carro.get('numero_renavam')
                dados_carro['ano_veiculo'] = dados_carro.get('ano_veiculo')

                dados_cliente = run_query("SELECT * FROM clientes WHERE id=?", (cliente_id,), fetch=True).iloc[
                    0].to_dict()
                # Recarrega a reserva para ter todos os campos
                dados_reserva = run_query("SELECT * FROM reservas WHERE id=?", (id_reserva_sel,), fetch=True).iloc[0]

                # --- VALIDAÇÃO DE CNH ---
                validade_cnh_db = dados_cliente.get('validade_cnh')

                if validade_cnh_db:
                    # Converte para objeto date. Se for string (como 'YYYY-MM-DD'), funciona.
                    validade_date = pd.to_datetime(validade_cnh_db).date()
                    if validade_date < date.today():
                        st.error(f"❌ CNH Expirada! Validade: {validade_date.strftime('%d/%m/%Y')}. Locação Bloqueada.")
                        st.stop()  # Bloqueia o restante da execução
                    else:
                        st.success(f"✅ CNH Válida! Expira em {validade_date.strftime('%d/%m/%Y')}.")
                else:
                    # Se o campo estiver vazio ou for None
                    st.warning("⚠️ Data de validade da CNH não encontrada no cadastro do cliente.")
                # --- FIM VALIDAÇÃO DE CNH ---

                st.markdown("---")

                # Recarrega a reserva para ter todos os campos
                dados_reserva = run_query("SELECT * FROM reservas WHERE id=?", (id_reserva_sel,), fetch=True).iloc[0]

                st.markdown("---")

                # 1. Confirmação de KM e Data de Saída
                st.subheader(f"Confirmar entrega de {dados_carro['modelo']} para {dados_cliente['nome']}")

                # --- INÍCIO DO FORMULÁRIO ---
                with st.form("entrega_form"):
                    col_km, col_data = st.columns(2)

                    # KM Saída Atual
                    km_saida_atual = int(dados_reserva['km_saida'])
                    km_confirma = col_km.number_input(
                        f"KM de Saída (Registrado na Reserva: {km_saida_atual})",
                        min_value=km_saida_atual,
                        value=km_saida_atual,
                        help="Confirme ou atualize a quilometragem exata no momento da saída."
                    )

                    # Usa a data de início da reserva como valor padrão
                    data_inicio_do_db = dados_reserva['data_inicio']

                    if isinstance(data_inicio_do_db, datetime):
                        data_prevista = data_inicio_do_db.date()
                    else:
                        data_prevista = data_inicio_do_db

                    # data_prevista agora é o objeto date correto para o st.date_input
                    data_saida = col_data.date_input("Data Real da Saída", data_prevista)

                    # Apenas o botão de submissão do formulário
                    submit_entrega = st.form_submit_button("✅ Finalizar Entrega e Gerar Contrato", type="primary")

                    if submit_entrega:
                        # 2. Atualiza Status do Carro para 'Locado'
                        run_query("UPDATE carros SET status='Locado', km_atual=? WHERE id=?",
                                  (km_confirma, carro_id))

                        # 3. Atualiza KM de Saída Real, a Data de Início na reserva
                        run_query("UPDATE reservas SET km_saida=?, data_inicio=?, reserva_status='Locada' WHERE id=?",
                                  (km_confirma, data_saida, id_reserva_sel))

                        st.toast("Entrega Confirmada! Carro Locado.", icon="🔑")
                        st.success("Entrega finalizada! O veículo agora está **Locado**. Baixe o contrato.")

                        # 4. Geração do Contrato (AGORA SALVANDO NO SESSION STATE)
                        pdf_bytes = gerar_contrato_pdf(
                            dados_cliente,
                            dados_carro,
                            data_saida,  # Usar a data real da saída
                            pd.to_datetime(dados_reserva['data_fim']).date()  # Data fim da reserva
                        )

                        # Salva o PDF no Session State
                        st.session_state.pdf_para_download = pdf_bytes
                        st.session_state.pdf_file_name = f"contrato_{dados_cliente['nome']}_{dados_carro['placa']}.pdf"

                        st.balloons()
                        #st.rerun() 

    # --- FORA DO FORMULÁRIO: EXIBIÇÃO DO DOWNLOAD BUTTON ---
    # Verifica se há um PDF no Session State para download
    if st.session_state.pdf_para_download is not None:
        st.markdown("---")
        st.download_button(
            label="📄 Baixar Contrato em PDF",
            data=st.session_state.pdf_para_download,
            file_name=st.session_state.pdf_file_name,
            mime="application/pdf"
        )
        # Limpa o estado após exibir o botão, para que ele não apareça em re-runs subsequentes
        st.session_state.pdf_para_download = None
        st.session_state.pdf_file_name = None
        #st.rerun()  # Opcional: Força um rerun final para limpar a UI

    else:
        st.info("Nenhuma reserva pendente de entrega. Verifique o menu '1. Reservar Veículo'.")


# 6. DEVOLUÇÃO
elif menu == "Devolução":
    st.title("🔄 Devolução de Veículo")

    # Inicializa o Session State para o PDF, se necessário
    if 'pdf_para_download' not in st.session_state:
        st.session_state.pdf_para_download = None
        st.session_state.pdf_file_name = None

    # MUDANÇA NO FILTRO: Busca reservas com reserva_status='Locada' (Carro em uso pelo cliente)
    query_dev = """
        SELECT 
            r.id, cl.nome, c.modelo, c.placa, r.km_saida, c.preco_km, c.diaria, 
            r.data_inicio, r.carro_id, r.cliente_id, r.km_franquia, r.adiantamento, 
            r.valor_multas, r.valor_danos, r.valor_outros 
        FROM reservas r 
        JOIN carros c ON r.carro_id = c.id 
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status='Ativa' AND r.reserva_status='Locada'
    """
    ativas = run_query(query_dev, fetch=True)

    if isinstance(ativas, str):
        st.error(f"Erro no banco de dados: {ativas}")
    elif not ativas.empty:
        opcoes = ativas.apply(lambda x: f"{x['id']} - {x['nome']} ({x['modelo']} - {x['placa']})", axis=1)
        opcoes_com_placeholder = ["Selecione a locação pendente..."] + opcoes.tolist()

        sel = st.selectbox("Selecione a Locação Pendente", opcoes_com_placeholder)

        if sel != "Selecione a locação pendente...":

            try:
                id_reserva_sel = int(sel.split(" - ")[0])
                reserva = ativas[ativas['id'] == id_reserva_sel].iloc[0]
            except:
                st.warning("Erro ao processar ID da reserva. Selecione novamente.")
                reserva = None

            if reserva is not None:

                try:
                    km_saida_safe = int(reserva['km_saida'])
                except:
                    km_saida_safe = 0
                    st.warning("Aviso: KM de saída com formato inválido no banco. Usando 0.")

                # Buscar dados completos do cliente para o recibo
                cliente_id = int(reserva['cliente_id'])
                dados_cliente = run_query("SELECT * FROM clientes WHERE id=?", (cliente_id,), fetch=True).iloc[0].to_dict()

                st.markdown("---")

                col_input1, col_input2, col_input3 = st.columns(3)

                with col_input1:
                    st.info(f"📍 KM de Saída: **{km_saida_safe}**")
                    km_volta = st.number_input(
                        "KM de Devolução (Atual)",
                        min_value=km_saida_safe,
                        value=km_saida_safe,
                        help="A quilometragem não pode ser menor que a da saída."
                    )

                with col_input2:
                    st.write("💧 **Serviços Extras**")
                    cobrar_lavagem = st.checkbox("Cobrar Lavagem?", value=False)
                    valor_lavagem = 0.0
                    if cobrar_lavagem:
                        valor_lavagem = st.number_input("Valor da Lavagem (R$)", value=50.0, step=5.0, key="dev_valor_lavagem")
                
                with col_input3:
                    st.write("💸 **Outros Custos**")
                    valor_multas = st.number_input("Valor de Multas (R$)", min_value=0.0, value=0.0, step=10.0, format="%.2f", key="dev_valor_multas")
                    valor_danos = st.number_input("Valor de Danos ao Veículo (R$)", min_value=0.0, value=0.0, step=10.0, format="%.2f", key="dev_valor_danos")
                    valor_outros = st.number_input("Outros Custos (R$)", min_value=0.0, value=0.0, step=10.0, format="%.2f", key="dev_valor_outros")

                # CÁLCULOS
                data_saida_real = pd.to_datetime(reserva['data_inicio']).date()
                data_devolucao = date.today()
                dias = (data_devolucao - data_saida_real).days
                dias_cobranca = max(dias, 1)

                km_rodados_totais = km_volta - km_saida_safe
                km_franquia_reserva = reserva['km_franquia'] if reserva['km_franquia'] is not None else 0
                if km_rodados_totais > km_franquia_reserva:
                    km_franquia_reserva = 0
                km_a_cobrar = max(0, km_rodados_totais - km_franquia_reserva)

                custo_km = km_a_cobrar * reserva['preco_km']
                custo_diarias = reserva['diaria'] * dias_cobranca

                subtotal_sem_adiantamento = custo_diarias + custo_km + valor_lavagem + valor_multas + valor_danos + valor_outros
                total_final = subtotal_sem_adiantamento - reserva['adiantamento']
                
                # Define o que será exibido como "Total a Pagar"
                if total_final >= 0:
                    label_total = "Total a Pagar (R$)"
                    valor_display = formatar_moeda(total_final)
                else:
                    label_total = "Valor a Receber (R$)"
                    valor_display = formatar_moeda(abs(total_final))

                st.markdown("---")
                st.subheader("Fatura de Locação")

                col_fat1, col_fat2, col_fat3, col_fat4 = st.columns(4)
                col_fat1.metric("Diárias Cobradas", f"{dias_cobranca} dias",
                                f"Início: {data_saida_real.strftime('%d/%m/%Y')}")
                col_fat2.metric("Custo Diárias", formatar_moeda(custo_diarias))
                col_fat3.metric("KM Rodados (Totais)", f"{km_rodados_totais} km")
                col_fat4.metric("KM a Cobrar (Após Franquia)", f"{km_a_cobrar} km")

                st.metric("KM de Franquia Aplicado", f"{km_franquia_reserva} km")
                st.metric("Adiantamento Recebido", formatar_moeda(reserva['adiantamento'] if reserva['adiantamento'] is not None else 0.0))
                
                # Novos custos extras exibidos
                if valor_lavagem > 0: st.metric("Custo Lavagem", formatar_moeda(valor_lavagem))
                if valor_multas > 0: st.metric("Custo Multas", formatar_moeda(valor_multas))
                if valor_danos > 0: st.metric("Custo Danos", formatar_moeda(valor_danos))
                if valor_outros > 0: st.metric("Outros Custos", formatar_moeda(valor_outros))

                st.markdown("---")
                st.metric(label_total, valor_display,
                                          help=f"Diárias ({formatar_moeda(custo_diarias)}) + KM ({formatar_moeda(custo_km)}) + Lavagem ({formatar_moeda(valor_lavagem)}) + Multas ({formatar_moeda(valor_multas)}) + Danos ({formatar_moeda(valor_danos)}) + Outros ({formatar_moeda(valor_outros)}) - Adiantamento ({formatar_moeda(reserva['adiantamento'] if reserva['adiantamento'] is not None else 0.0)})")

                if st.button("✅ Finalizar Devolução e Liberar Carro", type="primary"):
                    # 1. Atualiza Status da Reserva e armazena os valores finais
                    # MUDANÇA: Altera o reserva_status para 'Finalizada' e salva novos custos
                    run_query("""
                        UPDATE reservas 
                        SET status='Finalizada', reserva_status='Finalizada', km_volta=?, custo_lavagem=?, valor_total=?, 
                        valor_multas=?, valor_danos=?, valor_outros=? 
                        WHERE id=?
                    """, (km_volta, valor_lavagem, subtotal_sem_adiantamento, 
                         valor_multas, valor_danos, valor_outros, int(reserva['id'])))

                    # 2. Atualiza Status do Carro para 'Disponível' e KM
                    run_query("UPDATE carros SET status='Disponível', km_atual=? WHERE id=?",
                              (km_volta, int(reserva['carro_id'])))

                    st.toast("Devolução Finalizada!", icon="🎉")
                    st.success(
                        f"Devolução da placa {reserva['placa']} finalizada. Total: {valor_display}. O veículo está novamente disponível.")

                    # Re-buscar dados completos do carro para o recibo
                    dados_carro_recibo = run_query("SELECT * FROM carros WHERE id=?", (int(reserva['carro_id']),), fetch=True).iloc[0].to_dict()
                    dados_carro_recibo['chassi'] = dados_carro_recibo.get('numero_chassi')
                    dados_carro_recibo['renavam'] = dados_carro_recibo.get('numero_renavam')

                    # Geração do Recibo em PDF
                    recibo_pdf_bytes = gerar_recibo_pdf(
                        dados_cliente,
                        dados_carro_recibo,
                        {
                            'data_inicio': data_saida_real,
                            'data_fim': data_devolucao,
                            'km_saida': km_saida_safe,
                            'km_volta': km_volta,
                            'km_franquia': km_franquia_reserva,
                            'dias_cobranca': dias_cobranca,
                            'custo_diarias': custo_diarias,
                            'custo_km': custo_km,
                            'valor_lavagem': valor_lavagem,
                            'valor_multas': valor_multas,
                            'valor_danos': valor_danos,
                            'valor_outros': valor_outros,
                            'adiantamento': reserva['adiantamento'] if reserva['adiantamento'] is not None else 0.0,
                            'total_final': total_final
                        }
                    )

                    # Salvar o recibo em Session State para download
                    st.session_state.pdf_para_download = recibo_pdf_bytes
                    st.session_state.pdf_file_name = f"recibo_{dados_cliente['nome']}_{reserva['placa']}_{date.today().strftime('%Y%m%d')}.pdf"

                    st.balloons()
                    #st.rerun()

            if st.session_state.pdf_para_download:
                st.download_button(
                    label="📥 Baixar Comprovante de Devolução",
                    data=st.session_state.pdf_para_download,
                    file_name=st.session_state.pdf_file_name,
                    mime="application/pdf",
                    key="download_recibo_pos_devolucao"
                )

    else:
        st.info("Nenhum veículo em locação para ser devolvido.")

# 7. HISTÓRICO
elif menu == "Histórico":
    st.title("📜 Histórico de Locações Finalizadas")

    st.subheader("Faturamento Mensal")

    # Obter lista de meses com locações finalizadas
    meses_db = run_query_dataframe(
        "SELECT DISTINCT strftime('%Y-%m', data_fim) AS mes FROM reservas WHERE reserva_status='Finalizada' ORDER BY mes DESC"
    )

    if not meses_db.empty:
        lista_meses = meses_db['mes'].tolist()
        # Define o mês atual como padrão, se disponível na lista, caso contrário, usa o primeiro da lista
        hoje_mes_str = date.today().strftime('%Y-%m')
        if hoje_mes_str in lista_meses:
            default_index = lista_meses.index(hoje_mes_str)
        else:
            default_index = 0 # Se o mês atual não está na lista, seleciona o primeiro

        mes_selecionado_str = st.selectbox(
            "Selecione o Mês", 
            lista_meses, 
            index=default_index, 
            format_func=lambda x: datetime.strptime(x, '%Y-%m').strftime('%B/%Y')
        )

        primeiro_dia_mes = datetime.strptime(mes_selecionado_str, '%Y-%m').date().replace(day=1)
        # Calcula o último dia do mês corretamente
        if primeiro_dia_mes.month == 12:
            ultimo_dia_mes = primeiro_dia_mes.replace(day=31)
        else:
            ultimo_dia_mes = (primeiro_dia_mes.replace(month=primeiro_dia_mes.month + 1, day=1) - timedelta(days=1))

        query_historico = f"""
            SELECT 
                r.id AS Reserva_ID, 
                cl.nome AS Cliente, 
                c.modelo AS Veiculo, 
                c.placa AS Placa, 
                r.data_inicio AS Inicio, 
                r.data_fim AS Fim, 
                r.valor_total AS Total_Faturado,
                -- Tratamento para evitar divisão por zero: se km_rodados for 0, retorna 0, senão calcula
                CASE WHEN (r.km_volta - r.km_saida) = 0 THEN 0 ELSE (r.valor_total / (r.km_volta - r.km_saida)) END AS Lucro_por_km,
                (r.km_volta - r.km_saida) AS Km_Rodados
            FROM reservas r
            JOIN clientes cl ON r.cliente_id = cl.id
            JOIN carros c ON r.carro_id = c.id
            WHERE r.reserva_status = 'Finalizada'
            AND r.data_fim BETWEEN '{primeiro_dia_mes}' AND '{ultimo_dia_mes}'
        """
        df_historico = run_query_dataframe(query_historico)

        if not df_historico.empty:
            faturamento_total = df_historico['Total_Faturado'].sum()
            st.metric(f"Faturamento Total em {datetime.strptime(mes_selecionado_str, '%Y-%m').strftime('%B/%Y')}", formatar_moeda(faturamento_total))

            # --- GRÁFICOS ---
            st.subheader("Análise de Desempenho")

            # Gráfico 1: Faturamento por Veículo
            faturamento_por_veiculo = df_historico.groupby('Veiculo')['Total_Faturado'].sum().sort_values(
                ascending=False)

            fig, ax = plt.subplots(figsize=(10, 6))
            faturamento_por_veiculo.plot(kind='bar', ax=ax, color='skyblue')
            ax.set_title(f"Faturamento Total por Modelo ({datetime.strptime(mes_selecionado_str, '%Y-%m').strftime('%B/%Y')})")
            ax.set_ylabel('Faturamento (R$)')
            ax.set_xlabel('Modelo do Veículo')
            plt.xticks(rotation=45, ha='right')
            plt.tight_layout()

            col_grafico1, col_grafico2 = st.columns(2)

            with col_grafico1:
                st.markdown("##### 📈 Faturamento por Modelo")
                st.pyplot(fig)

            # Gráfico 2: Participação no Faturamento (Gráfico de Pizza)
            fig_pizza, ax_pizza = plt.subplots(figsize=(8, 8))
            ax_pizza.pie(faturamento_por_veiculo, labels=faturamento_por_veiculo.index, autopct='%1.1f%%',
                         startangle=90, colors=plt.cm.Paired.colors)
            ax_pizza.set_title(f"Participação de Mercado ({datetime.strptime(mes_selecionado_str, '%Y-%m').strftime('%B/%Y')})")

            with col_grafico2:
                st.markdown("##### 🍕 Participação no Faturamento")
                st.pyplot(fig_pizza)

            st.markdown("---")
            st.subheader("Detalhes das Locações Finalizadas")

            # --- BACKUP E DOWNLOAD CSV ---
            df_historico_display = df_historico.copy()

            # Formata colunas para exibição na tela e para o CSV
            df_historico_display['Inicio'] = pd.to_datetime(df_historico_display['Inicio']).dt.strftime('%d/%m/%Y')
            df_historico_display['Fim'] = pd.to_datetime(df_historico_display['Fim']).dt.strftime('%d/%m/%Y')
            df_historico_display['Total_Faturado'] = df_historico_display['Total_Faturado'].apply(formatar_moeda)
            df_historico_display['Lucro_por_km'] = df_historico_display['Lucro_por_km'].apply(formatar_moeda)

            csv = df_historico_display.to_csv(index=False, sep=';').encode('utf-8')

            st.download_button(
                label="💾 Baixar Histórico Completo (CSV)",
                data=csv,
                file_name=f'historico_locacoes_{mes_selecionado_str}.csv',
                mime='text/csv',
                key='download-csv'
            )

            st.dataframe(
                df_historico_display[['Cliente', 'Veiculo', 'Placa', 'Inicio', 'Fim', 'Total_Faturado', 'Km_Rodados', 'Lucro_por_km']],
                width='stretch'
            )

        else:
            st.info("Nenhuma locação finalizada para o mês selecionado ou dados insuficientes para gerar gráficos.")

    else:
        st.info("Nenhuma locação finalizada no histórico.")

# 8. RELATÓRIOS (NOVA ABA)
elif menu == "Relatórios":
    st.title("📈 Relatórios")
    st.write("Aqui você pode gerar relatórios de disponibilidade da frota.")

    # Seleção de Mês e Ano
    col_mes, col_ano = st.columns(2)
    mes_selecionado = col_mes.selectbox("Selecione o Mês", range(1, 13), index=date.today().month - 1,
                                       format_func=lambda x: datetime(2000, x, 1).strftime('%B'))
    ano_selecionado = col_ano.selectbox("Selecione o Ano", range(datetime.now().year - 2, datetime.now().year + 3),
                                       index=2)

    # Obtém o primeiro e último dia do mês selecionado
    primeiro_dia_mes = date(ano_selecionado, mes_selecionado, 1)
    if mes_selecionado == 12:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado, 31)
    else:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado + 1, 1) - timedelta(days=1)

    # Dataframe de carros ativos (não excluídos)
    df_carros = run_query_dataframe(f"SELECT id, modelo, placa FROM carros WHERE status != '{STATUS_CARRO['EXCLUIDO']}'")

    if df_carros.empty:
        st.warning("Nenhum veículo ativo encontrado para gerar o relatório.")
    else:
        # Obter todas as reservas ativas (Locada ou Reservada) para o período do mês
        query_reservas = f"""
            SELECT carro_id, data_inicio, data_fim, reserva_status
            FROM reservas
            WHERE (reserva_status = 'Reservada' OR reserva_status = 'Locada')
            AND (data_inicio <= '{ultimo_dia_mes}' AND data_fim >= '{primeiro_dia_mes}')
        """
        df_reservas = run_query_dataframe(query_reservas)

        # Criar a estrutura para o relatório
        # A primeira coluna será o nome do veículo, as outras serão os dias do mês
        dias_no_mes = (ultimo_dia_mes - primeiro_dia_mes).days + 1
        colunas_dias = [f"{d:02d}" for d in range(1, dias_no_mes + 1)]

        # Inicializa o DataFrame do relatório com a coluna de veículos
        df_relatorio_data = {'Veículo': df_carros['modelo'] + " (" + df_carros['placa'] + ")"}
        for dia in colunas_dias:
            df_relatorio_data[dia] = '' # Preenche com vazio inicialmente

        df_relatorio = pd.DataFrame(df_relatorio_data)

        # Mapear IDs de carro para índice no df_relatorio para atualização eficiente
        carro_id_to_index = {carro_id: i for i, carro_id in enumerate(df_carros['id'])}        

        # Preencher o DataFrame do relatório com base nas reservas
        for index, row in df_reservas.iterrows():
            carro_id = row['carro_id']
            reserva_status = row['reserva_status']
            data_inicio = pd.to_datetime(row['data_inicio']).date()
            data_fim = pd.to_datetime(row['data_fim']).date()

            if carro_id in carro_id_to_index:
                idx_df_relatorio = carro_id_to_index[carro_id]
                
                for d in range(1, dias_no_mes + 1):
                    dia_atual = date(ano_selecionado, mes_selecionado, d)
                    if data_inicio <= dia_atual <= data_fim:
                        col_dia = f"{d:02d}"
                        if reserva_status == 'Reservada':
                            df_relatorio.at[idx_df_relatorio, col_dia] = 'Reservado'
                        elif reserva_status == 'Locada':
                            df_relatorio.at[idx_df_relatorio, col_dia] = 'Locado'

        # Após preencher os status de reservas/locações, preencher o restante como 'Disponível'
        for r_idx in range(len(df_relatorio)):
            for c_idx in colunas_dias:
                if df_relatorio.at[r_idx, c_idx] == '':
                    df_relatorio.at[r_idx, c_idx] = 'Disponível'

        st.dataframe(df_relatorio, hide_index=True)

        st.markdown("---")
        st.subheader("Gerar Relatório Excel")

        if st.button("📊 Gerar e Baixar Relatório de Disponibilidade", type="primary"):
            output = io.BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"Disponibilidade {mes_selecionado:02d}-{ano_selecionado}"

            # Obter nomes dos veículos (serão os cabeçalhos das colunas do Excel, a partir da coluna B)
            # Usa a coluna 'Veículo' do df_relatorio que já inclui Modelo e Placa
            vehicle_names_with_plate = df_relatorio['Veículo'].tolist()
            
            # Obter números dos dias (serão os cabeçalhos das linhas do Excel, a partir da linha 2)
            day_numbers_str = colunas_dias # ex: ['01', '02', ...]

            # Célula A1 vazia ou com um rótulo
            sheet.cell(row=1, column=1, value="Dia/Veículo")

            # Escrever nomes dos veículos como cabeçalhos de coluna (linha 1, começando da coluna B)
            for col_idx, vehicle_name in enumerate(vehicle_names_with_plate, start=2):
                sheet.cell(row=1, column=col_idx, value=vehicle_name)

            # Escrever números dos dias como cabeçalhos de linha (coluna 1, começando da linha 2)
            for row_idx, day_str in enumerate(day_numbers_str, start=2):
                sheet.cell(row=row_idx, column=1, value=int(day_str)) # Converte para int para exibição

            # Estilo para o cabeçalho
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            header_font = Font(bold=True)
            
            # Aplicar à primeira linha (cabeçalhos dos veículos)
            for col in range(1, len(vehicle_names_with_plate) + 2):
                sheet.cell(row=1, column=col).fill = header_fill
                sheet.cell(row=1, column=col).font = header_font
            
            # Aplicar à primeira coluna (cabeçalhos dos dias)
            for row in range(1, len(day_numbers_str) + 2):
                sheet.cell(row=row, column=1).fill = header_fill
                sheet.cell(row=row, column=1).font = header_font

            # Preencher dados e aplicar formatação condicional
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
            orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Laranja
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Vermelho

            # Preencher as células de dados (status)
            # Iterar pelos dias (que agora são as linhas no Excel)
            for day_excel_row_idx, day_str in enumerate(day_numbers_str, start=2):
                # Iterar pelos veículos (que agora são as colunas no Excel)
                for vehicle_excel_col_idx, vehicle_name_full in enumerate(vehicle_names_with_plate, start=2):
                    
                    # Para obter o status de um dia e veículo específicos,
                    # precisamos encontrar a linha no df_relatorio que corresponde ao vehicle_name_full
                    # e então pegar o valor da coluna 'day_str'.
                    
                    # Encontrar o índice da linha do veículo no df_relatorio original
                    original_df_row_index = df_relatorio[df_relatorio['Veículo'] == vehicle_name_full].index[0]
                    
                    # Obter o status para o dia específico dessa linha
                    status = df_relatorio.at[original_df_row_index, day_str]
                    
                    cell = sheet.cell(row=day_excel_row_idx, column=vehicle_excel_col_idx, value=status)

                    if status == 'Disponível':
                        cell.fill = green_fill
                    elif status == 'Reservado':
                        cell.fill = orange_fill
                    elif status == 'Locado':
                        cell.fill = red_fill

            # Ajustar largura das colunas
            sheet.column_dimensions['A'].width = 10 # Largura para a coluna dos dias
            for col_idx in range(2, len(vehicle_names_with_plate) + 2):
                sheet.column_dimensions[chr(64 + col_idx)].width = 25 # Largura para os nomes dos veículos

            workbook.save(output)
            output.seek(0)

            st.download_button(
                label="Download Excel",
                data=output.getvalue(),
                file_name=f"relatorio_disponibilidade_{mes_selecionado:02d}-{ano_selecionado}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Relatório Excel gerado com sucesso!")

# 8. GERENCIAR USUÁRIOS (APENAS ADMIN)
elif menu == "👥 Gerenciar Usuários":
    st.title("👥 Gerenciamento de Usuários")

    if not check_permission('manage_users'):
        st.error("❌ Você não tem permissão para acessar esta seção.")
        st.stop()

    tab_listar, tab_criar, tab_auditoria = st.tabs(["Listar Usuários", "Criar Usuário", "Logs de Auditoria"])

    with tab_listar:
        st.subheader("Usuários Cadastrados")

        users = auth_manager.get_users()

        if users:
            # Formatar dados para exibição
            df_users = pd.DataFrame(users)
            df_users['status'] = df_users['is_active'].map({True: 'Ativo', False: 'Inativo'})
            df_users['role_display'] = df_users['role'].map({
                'admin': 'Administrador',
                'manager': 'Gerente',
                'employee': 'Funcionário',
                'viewer': 'Visualizador'
            })

            st.dataframe(
                df_users[['username', 'full_name', 'email', 'role_display', 'status', 'created_at', 'last_login']],
                column_config={
                    'username': 'Usuário',
                    'full_name': 'Nome Completo',
                    'email': 'Email',
                    'role_display': 'Nível',
                    'status': 'Status',
                    'created_at': 'Criado em',
                    'last_login': 'Último Login'
                },
                use_container_width=True
            )

            # Seleção de usuário para edição
            st.markdown("---")
            st.subheader("Editar Usuário")

            user_options = [f"{u['id']} - {u['username']} ({u['full_name']})" for u in users]
            selected_user = st.selectbox("Selecione usuário para editar", ["Nenhum"] + user_options)

            if selected_user != "Nenhum":
                user_id = int(selected_user.split(" - ")[0])
                user_data = next((u for u in users if u['id'] == user_id), None)

                if user_data:
                    st.markdown(f"**Editando:** {user_data['username']}")

                    with st.form(f"edit_user_{user_id}"):
                        col1, col2 = st.columns(2)

                        with col1:
                            new_full_name = st.text_input("Nome Completo", value=user_data['full_name'] or "")
                            new_email = st.text_input("Email", value=user_data['email'] or "")

                        with col2:
                            new_role = st.selectbox(
                                "Nível de Acesso",
                                options=list(USER_ROLES.keys()),
                                format_func=lambda x: USER_ROLES[x],
                                index=list(USER_ROLES.keys()).index(user_data['role'])
                            )
                            new_active = st.checkbox("Usuário Ativo", value=user_data['is_active'])

                        new_password = st.text_input("Nova Senha (deixe vazio para manter)", type="password")

                        if st.form_submit_button("💾 Salvar Alterações", type="primary"):
                            updates = {
                                'full_name': new_full_name,
                                'email': new_email,
                                'role': new_role,
                                'is_active': new_active
                            }

                            if new_password:
                                updates['password'] = new_password

                            success, message = auth_manager.update_user(user_id, updates)

                            if success:
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)

                    # Botão para remover usuário
                    if st.button("🗑️ Desativar Usuário", type="secondary"):
                        if user_data['username'] == current_user['username']:
                            st.error("❌ Você não pode desativar seu próprio usuário!")
                        else:
                            success, message = auth_manager.delete_user(user_id)
                            if success:
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)
        else:
            st.info("Nenhum usuário cadastrado.")

    with tab_criar:
        st.subheader("Criar Novo Usuário")

        with st.form("create_user_form"):
            col1, col2 = st.columns(2)

            with col1:
                username = st.text_input("Nome de Usuário", help="Nome único para login")
                full_name = st.text_input("Nome Completo")
                email = st.text_input("Email")

            with col2:
                password = st.text_input("Senha", type="password", help="Mínimo 6 caracteres")
                confirm_password = st.text_input("Confirmar Senha", type="password")
                role = st.selectbox(
                    "Nível de Acesso",
                    options=list(USER_ROLES.keys()),
                    format_func=lambda x: USER_ROLES[x]
                )

            if st.form_submit_button("👤 Criar Usuário", type="primary"):
                if not username or not password or not full_name:
                    st.error("❌ Preencha todos os campos obrigatórios!")
                elif password != confirm_password:
                    st.error("❌ As senhas não coincidem!")
                elif len(password) < 6:
                    st.error("❌ A senha deve ter pelo menos 6 caracteres!")
                else:
                    success, message = auth_manager.create_user(
                        username, password, role, full_name, email
                    )

                    if success:
                        st.success(message)
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(message)

    with tab_auditoria:
        st.subheader("Logs de Auditoria")

        logs = auth_manager.get_audit_logs(200)  # Últimos 200 registros

        if logs:
            df_logs = pd.DataFrame(logs)
            df_logs['timestamp'] = pd.to_datetime(df_logs['timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')

            st.dataframe(
                df_logs,
                column_config={
                    'timestamp': 'Data/Hora',
                    'username': 'Usuário',
                    'action': 'Ação',
                    'resource': 'Recurso',
                    'details': 'Detalhes',
                    'ip_address': 'IP'
                },
                use_container_width=True
            )

            # Resumo de atividades
            st.markdown("---")
            st.subheader("📊 Resumo de Atividades")

            col1, col2, col3 = st.columns(3)

            with col1:
                total_logins = len([l for l in logs if l['action'] == 'login'])
                st.metric("Total de Logins", total_logins)

            with col2:
                total_users_created = len([l for l in logs if l['action'] == 'user_created'])
                st.metric("Usuários Criados", total_users_created)

            with col3:
                active_users = len(set([l['username'] for l in logs if l['username']]))
                st.metric("Usuários Ativos", active_users)
        else:
            st.info("Nenhum log de auditoria encontrado.")

# 9. BACKUP (NOVA ABA)
elif menu == "Backup":
    interface_backup()
